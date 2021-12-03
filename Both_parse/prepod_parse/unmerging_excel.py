#!/usr/bin/env python
# coding: utf-8

# # IMPORT

# In[1]:


import pandas as pd
import numpy as np

import traceback

import os
import sys
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl import Workbook
from openpyxl import load_workbook

pd.set_option('display.max_rows', None)


# # Unmerging cells

# In[2]:

wbook=openpyxl.load_workbook(os.path.join(sys.path[0], r"rasp\mer", "eht_mer.xlsx"))

def de_merge():
    c = 0
    for cell_group in merged_renges:
        c += 1
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value
                    
    if c != 0:
        de_merge()

num = -1
for sheet in wbook.worksheets:
    num += 1
    if sheet.sheet_state == 'hidden':
        wbook.remove_sheet(sheet)
    else:
        merged_renges = sheet.merged_cells.ranges
        de_merge()

wbook.save(os.path.join(sys.path[0], r"rasp\unmer", "eht_unmer.xlsx"))


# In[3]:


wbook=openpyxl.load_workbook(os.path.join(sys.path[0], r"rasp\mer", "eiu_mer.xlsx"))

def de_merge():
    c = 0
    for cell_group in merged_renges:
        c += 1
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value
                    
    if c != 0:
        de_merge()

num = -1
for sheet in wbook.worksheets:
    num += 1
    if sheet.sheet_state == 'hidden':
        wbook.remove_sheet(sheet)
    else:
        merged_renges = sheet.merged_cells.ranges
        de_merge()

wbook.save(os.path.join(sys.path[0], r"rasp\unmer", "eiu_unmer.xlsx"))


# In[4]:

wbook=openpyxl.load_workbook(os.path.join(sys.path[0], r"rasp\mer", "pma_mer.xlsx"))

def de_merge():
    c = 0
    for cell_group in merged_renges:
        c += 1
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value
                    
    if c != 0:
        de_merge()

num = -1
for sheet in wbook.worksheets:
    num += 1
    if sheet.sheet_state == 'hidden':
        wbook.remove_sheet(sheet)
    else:
        merged_renges = sheet.merged_cells.ranges
        de_merge()

wbook.save(os.path.join(sys.path[0], r"rasp\unmer", "pma_unmer.xlsx"))


# In[5]:

wbook=openpyxl.load_workbook(os.path.join(sys.path[0], r"rasp\mer", "uits_mer.xlsx"))


def de_merge():
    c = 0
    for cell_group in merged_renges:
        c += 1
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value
                    
    if c != 0:
        de_merge()

num = -1
for sheet in wbook.worksheets:
    num += 1
    if sheet.sheet_state == 'hidden':
        wbook.remove_sheet(sheet)
    else:
        merged_renges = sheet.merged_cells.ranges
        de_merge()

wbook.save(os.path.join(sys.path[0], r"rasp\unmer", "uits_unmer.xlsx"))


# In[6]:

wbook=openpyxl.load_workbook(os.path.join(sys.path[0], r"rasp\mer", "tf_mer.xlsx"))

def de_merge():
    c = 0
    for cell_group in merged_renges:
        c += 1
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = top_left_cell_value
                    
    if c != 0:
        de_merge()

num = -1
for sheet in wbook.worksheets:
    num += 1
    if sheet.sheet_state == 'hidden':
        wbook.remove_sheet(sheet)
    else:
        merged_renges = sheet.merged_cells.ranges
        de_merge()
wbook.save(os.path.join(sys.path[0], r"rasp\unmer", "tf_unmer.xlsx"))

